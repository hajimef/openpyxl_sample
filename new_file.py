from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws['A1'] = 100
ws.cell(row = 1, column = 2, value = 'abc')
wb.save('sample.xlsx')
