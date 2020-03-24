from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active
rng = ws['B2':'D3']
for row in rng:
    for c in row:
        c.value = 123
        c.fill = PatternFill(patternType = 'solid', fgColor = '00ffff') 
wb.save('fill.xlsx')

