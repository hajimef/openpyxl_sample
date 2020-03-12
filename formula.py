from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws['A1'] = 100
ws['A2'] = 200
ws['B1'] = 5
ws['B2'] = 3
ws['C1'] = '=A1*B1'
ws['C2'] = '=A2*B2'
ws['C3'] = '=SUM(C1:C2)'
wb.save('formula.xlsx')

