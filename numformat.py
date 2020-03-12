# coding: utf-8
from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active
ws['A1'] = 1234567
ws['A1'].number_format = '#,##0'
ws['A2'] = 0.123
ws['A2'].number_format = '0.00%'
ws['A3'] = datetime.datetime(2020, 1, 23, 12, 34, 56)
ws['A3'].number_format = u'yyyy年m月d日 hh時mm分ss秒' 
wb.save('numformat.xlsx')
