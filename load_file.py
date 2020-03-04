from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')
ws = wb.active
ws['A2'] = 'テスト'
wb.save('sample.xlsx')
